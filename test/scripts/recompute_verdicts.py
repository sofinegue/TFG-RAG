"""Recalcula la columna `veredicto_umbral` y los colores de fondo en los Excel
de evaluación a partir de las columnas `coincidencia_%` y `relevancia_%`.

Útil cuando solo cambia el umbral y no se quiere re-llamar al LLM/embeddings.

Uso:
    python -m test.scripts.recompute_verdicts            # umbral por defecto (EVAL_THRESHOLD o 65)
    python -m test.scripts.recompute_verdicts --threshold 70
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

REPO_ROOT = Path(__file__).resolve().parents[2]
EVAL_DIR = REPO_ROOT / "test" / "evaluation"

OK_LIGHT = PatternFill("solid", fgColor="C6EFCE")  # ambas >= umbral
OK_DARK  = PatternFill("solid", fgColor="63BE7B")  # ambas >= 75
KO_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # alguna >= umbral
KO_RED    = PatternFill("solid", fgColor="FFC7CE")  # ninguna >= umbral


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def recompute(path: Path, threshold: int) -> tuple[int, int]:
    wb = load_workbook(path)
    ok_total = ko_total = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        h = [c.value for c in ws[1]]
        try:
            ic = h.index("coincidencia_%")
            ir = h.index("relevancia_%")
            iv = h.index("veredicto_umbral")
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2):
            c = _f(row[ic].value)
            r = _f(row[ir].value)
            cell = row[iv]
            if c is None or r is None:
                cell.value = "KO"
                cell.fill = KO_RED
                ko_total += 1
                continue
            if c >= threshold and r >= threshold:
                cell.value = "OK"
                cell.fill = OK_DARK if (c >= 75 and r >= 75) else OK_LIGHT
                ok_total += 1
            else:
                cell.value = "KO"
                cell.fill = KO_YELLOW if (c >= threshold or r >= threshold) else KO_RED
                ko_total += 1
    wb.save(path)
    return ok_total, ko_total


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--threshold", type=int,
                        default=int(os.getenv("EVAL_THRESHOLD", "65")))
    args = parser.parse_args()
    print(f"Umbral aplicado: {args.threshold}")
    grand_ok = grand_ko = 0
    for xlsx in sorted(EVAL_DIR.glob("*.xlsx")):
        ok, ko = recompute(xlsx, args.threshold)
        total = ok + ko
        pct = (ok * 100 / total) if total else 0
        print(f"  {xlsx.name:>15} -> OK={ok}/{total} ({pct:.1f}%)")
        grand_ok += ok
        grand_ko += ko
    grand_total = grand_ok + grand_ko
    grand_pct = (grand_ok * 100 / grand_total) if grand_total else 0
    print(f"\n  {'TOTAL':>15} -> OK={grand_ok}/{grand_total} ({grand_pct:.1f}%)")


if __name__ == "__main__":
    main()
