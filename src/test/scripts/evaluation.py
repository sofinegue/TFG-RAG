"""
evaluation.py
=============

Fase 2 del testing del TFG: **evaluación** automática de los resultados
generados por :mod:`execute_tests`.

Uso desde línea de comandos:
    python -m test.scripts.evaluation
    python -m test.scripts.evaluation --use-case cvs
    python -m test.scripts.evaluation --batch-size 20 --max-workers 8

Argumentos:
    --use-case      Caso de uso: cvs | eu | wiki (default: todos)
    --batch-size    Preguntas por lote LLM (default: 10)
    --max-workers   Threads paralelos (default: 4)
    --limit         Máximo nº de preguntas a evaluar

Funcionamiento
---------------
Por cada fichero ``test/results/results_<gold_standard>.json``:

* **Coincidencia (%)** – un LLM (por lotes de N preguntas) juzga cuánto
    cubre la respuesta real respecto a la esperada (entidades, hechos).
* **Relevancia (%)** – un LLM juzga si la respuesta real responde de forma
    coherente y pertinente a la pregunta planteada.
* **Veredicto OK/KO** – OK si coincidencia ≥ umbral **y** relevancia ≥ umbral.

El LLM se llama **en lotes** (por defecto 10 preguntas por llamada) para
minimizar coste y latencia, y devuelve ambas métricas en la misma respuesta.

Genera **un Excel por caso de uso** (``cvs.xlsx``, ``eu.xlsx``, ``wiki.xlsx``)
con **una hoja por idioma**. Columnas:

#. ``id``
#. ``categoria``
#. ``pregunta``
#. ``respuesta_esperada``
#. ``respuesta``
#. ``num_llamadas_llm``
#. ``tiempo_total_s``
#. ``tokens_in``
#. ``tokens_out``
#. ``coincidencia_%``
#. ``relevancia_%``
#. ``veredicto_umbral``     – OK si coincidencia ≥ 80 % **y** relevancia ≥ 80 %
#. ``justificacion``        – texto breve emitido por el LLM
#. ``error``                – mensaje de error si la consulta original falló

Uso
---
.. code-block:: powershell

    python -m test.scripts.evaluation
    python -m test.scripts.evaluation --model gpt-5-mini --threshold 80
    python -m test.scripts.evaluation --use-case cvs --batch-size 5

Dependencias
------------
``openpyxl`` para escribir los Excel.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, dict, Optional, tuple

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from openai import AzureOpenAI  # noqa: E402

from src.config import config, safe_create_kwargs  # noqa: E402  (carga el .env)
from test.scripts.pricing import token_cost_usd  # noqa: E402

# ---------------------------------------------------------------------------
# Configuración leída desde .env
# ---------------------------------------------------------------------------
def _resolve_path(env_var: str, default: str) -> Path:
    raw = os.getenv(env_var, default)
    p = Path(raw)
    return p if p.is_absolute() else (REPO_ROOT / p)


def _env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None or raw == "":
        return default
    try:
        return int(raw)
    except ValueError:
        return default


def _env_float(name: str, default: float) -> float:
    raw = os.getenv(name)
    if raw is None or raw == "":
        return default
    try:
        return float(raw)
    except ValueError:
        return default


RESULTS_DIR = _resolve_path("TEST_RESULTS_DIR", "src/test/results")
EVAL_DIR    = _resolve_path("TEST_EVAL_DIR",    "src/test/evaluation")
EVAL_DIR.mkdir(parents=True, exist_ok=True)

RESULTS_GLOB = os.getenv("TEST_RESULTS_GLOB", "results_gold_standard_*.json")

DEFAULT_MODEL     = os.getenv("CHAT_MODEL_EVAL") or os.getenv("EVAL_MODEL", "gpt-5-mini")
DEFAULT_THRESHOLD = _env_int("EVAL_THRESHOLD", 65)
DEFAULT_THRESHOLD_HIGH = _env_int("EVAL_THRESHOLD_HIGH", 75)
BATCH_SIZE        = _env_int("EVAL_BATCH_SIZE", 10)
MAX_RETRIES       = _env_int("EVAL_MAX_RETRIES", 2)
MAX_TOKENS_BATCH  = _env_int("EVAL_MAX_TOKENS", 2000)
TEMPERATURE       = _env_float("EVAL_TEMPERATURE", 0.0)
RETRY_BACKOFF_S   = _env_float("EVAL_RETRY_BACKOFF_S", 0.6)
JUSTIF_MAX_CHARS  = _env_int("EVAL_JUSTIFICACION_MAX_CHARS", 300)

_DEFAULT_SYSTEM_PROMPT = (
    "Eres un evaluador imparcial de sistemas RAG. Comparas respuestas de un "
    "sistema con las respuestas esperadas (gold standard) y con las preguntas "
    "originales. Juzgas cobertura factual y coherencia con la pregunta. "
    "Respondes SIEMPRE con un JSON array válido y nada más."
)
EVAL_SYSTEM_PROMPT = os.getenv("EVAL_SYSTEM_PROMPT", _DEFAULT_SYSTEM_PROMPT)


# ---------------------------------------------------------------------------
# Clientes (lazy)
# ---------------------------------------------------------------------------
_clients_cache: dict[str, tuple[AzureOpenAI, str]] = {}


def get_evaluator_client(model_name: str) -> tuple[AzureOpenAI, str]:
    """Devuelve ``(client, deployment)`` para el modelo evaluador."""
    if model_name in _clients_cache:
        return _clients_cache[model_name]
    cfg = config.get_model_config(model_name)
    client = AzureOpenAI(
        api_key=cfg.api_key,
        api_version=cfg.api_version,
        azure_endpoint=cfg.api_base,
    )
    _clients_cache[model_name] = (client, cfg.deployment)
    return client, cfg.deployment


# ---------------------------------------------------------------------------
# Prompt de evaluación por lotes
# ---------------------------------------------------------------------------
EVAL_BATCH_TEMPLATE = """Evalúa las siguientes {n} respuestas de un sistema RAG.
Para CADA una, asigna DOS puntuaciones:
1. coincidencia_pct: cuánto cubre la respuesta real respecto a la respuesta
   esperada (mismas entidades, mismos hechos, misma información).
2. relevancia_pct: hasta qué punto la respuesta real responde de forma
   coherente, directa y pertinente a la PREGUNTA planteada. Esta segunda
   puntuación se basa en la relación entre pregunta y respuesta real, NO en la
   similitud con la respuesta esperada.

{items}

Devuelve EXCLUSIVAMENTE un JSON array con {n} objetos (sin markdown, sin texto
adicional), en el MISMO ORDEN, con esta forma:
[
  {{"idx": 1, "coincidencia_pct": <0-100>, "relevancia_pct": <0-100>, "justificacion": "<máx 200 chars>"}},
  ...
]

Criterio para coincidencia_pct:
- 100: la respuesta real cubre toda la información de la esperada
- 80+: cubre la gran mayoría de entidades/hechos
- 50-79: cubre parcialmente
- <50: omite información clave o contiene errores graves

Criterio para relevancia_pct:
- 100: responde exactamente a la pregunta, sin desviaciones ni contradicciones
- 80+: responde bien a la pregunta, con pequeños matices mejorables
- 50-79: responde solo en parte, es ambigua o mezcla contenido secundario
- <50: responde poco o nada a la pregunta, o resulta incoherente/off-topic"""


def _format_batch_items(batch: list[dict[str, Any]]) -> str:
    parts = []
    for i, item in enumerate(batch, 1):
        esperada = item["esperada"]
        if isinstance(esperada, (list, dict)):
            esperada = json.dumps(esperada, ensure_ascii=False)
        parts.append(
            f"--- PREGUNTA {i} ---\n"
            f"Pregunta: {item['pregunta']}\n"
            f"Respuesta esperada: {esperada}\n"
            f"Respuesta real: {item['real']}\n"
        )
    return "\n".join(parts)


_JSON_ARRAY_RE = re.compile(r"\[[\s\S]*\]")


def _safe_json_array_extract(text: str) -> Optional[list[dict[str, Any]]]:
    if not text:
        return None
    text = text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if "\n" in text:
            text = text.split("\n", 1)[1]
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass
    m = _JSON_ARRAY_RE.search(text)
    if not m:
        return None
    try:
        data = json.loads(m.group(0))
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass
    return None


def _coerce_pct(value: Any) -> Optional[int]:
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    return max(0, min(100, int(round(v))))


# ---------------------------------------------------------------------------
# Evaluación por lotes (coincidencia + relevancia via LLM)
# ---------------------------------------------------------------------------
def evaluate_batch_metrics(
    batch: list[dict[str, Any]],
    model_name: str,
    max_retries: int = MAX_RETRIES,
) -> list[dict[str, Any]]:
    """Evalúa un lote de preguntas con una sola llamada LLM.

    Cada item en `batch` debe tener: pregunta, esperada, real.
    Devuelve lista de dicts con coincidencia_pct, relevancia_pct, justificacion,
    eval_error.
    """
    client, deployment = get_evaluator_client(model_name)
    n = len(batch)

    items_text = _format_batch_items(batch)
    user_prompt = EVAL_BATCH_TEMPLATE.format(n=n, items=items_text)

    last_error: Optional[str] = None
    last_tokens_in = 0
    last_tokens_out = 0
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                **safe_create_kwargs(
                    model=deployment,
                    messages=[
                        {"role": "system", "content": EVAL_SYSTEM_PROMPT},
                        {"role": "user",   "content": user_prompt},
                    ],
                    temperature=TEMPERATURE,
                    max_completion_tokens=MAX_TOKENS_BATCH,
                )
            )
            if response.usage:
                last_tokens_in  = response.usage.prompt_tokens or 0
                last_tokens_out = response.usage.completion_tokens or 0

            raw = response.choices[0].message.content or ""
            results = _safe_json_array_extract(raw)
            if not results or len(results) < n:
                last_error = f"JSON array inválido o incompleto: {raw[:200]}"
                continue

            tokens_in_per_q  = last_tokens_in  // n if n > 0 else 0
            tokens_out_per_q = last_tokens_out // n if n > 0 else 0
            parsed = []
            for i in range(n):
                item = results[i] if i < len(results) else {}
                parsed.append({
                    "coincidencia_pct": _coerce_pct(item.get("coincidencia_pct")),
                    "relevancia_pct": _coerce_pct(item.get("relevancia_pct")),
                    "justificacion": str(item.get("justificacion", ""))[:JUSTIF_MAX_CHARS],
                    "eval_error": None,
                    "tokens_eval_in":  tokens_in_per_q,
                    "tokens_eval_out": tokens_out_per_q,
                })
            return parsed

        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            time.sleep(RETRY_BACKOFF_S * attempt)

    # Fallback: devolver error para todo el batch (con tokens de la última llamada)
    tokens_in_per_q  = last_tokens_in  // n if n > 0 else 0
    tokens_out_per_q = last_tokens_out // n if n > 0 else 0
    return [
        {
            "coincidencia_pct": None, "relevancia_pct": None,
            "justificacion": "",
            "eval_error": last_error or "desconocido",
            "tokens_eval_in":  tokens_in_per_q,
            "tokens_eval_out": tokens_out_per_q,
        }
        for _ in range(n)
    ]


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
def _import_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        sys.exit(
            "❌ Falta openpyxl. Instálalo con `pip install openpyxl`.\n"
            f"   Detalle: {e}"
        )
    return Workbook, Font, PatternFill, Alignment


COLUMNS = [
    "id", "categoria", "pregunta", "respuesta_esperada", "respuesta",
    "num_llamadas_llm", "tiempo_total_s", "tokens_in", "tokens_out",
    "coincidencia_%", "relevancia_%", "veredicto_umbral",
    "justificacion",
    "coste_rag_usd", "coste_eval_usd", "coste_total_usd",
    "error",
]


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def write_excel_for_use_case(
    use_case: str,
    sheets: dict[str, list[dict[str, Any]]],
) -> Path:
    Workbook, Font, PatternFill, Alignment = _import_openpyxl()
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    # OK light green: ambas metricas >= 65 (umbral)
    ok_light_fill   = PatternFill("solid", fgColor="C6EFCE")
    # OK dark green: ambas metricas >= 75
    ok_dark_fill    = PatternFill("solid", fgColor="63BE7B")
    # KO yellow: al menos una metrica >= 65
    ko_yellow_fill  = PatternFill("solid", fgColor="FFEB9C")
    # KO red: ninguna metrica >= 65
    ko_red_fill     = PatternFill("solid", fgColor="FFC7CE")
    wrap            = Alignment(wrap_text=True, vertical="top")

    for lang in sorted(sheets):
        rows = sheets[lang]
        ws = wb.create_sheet(title=lang.upper()[:31] or "default")
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap

        for r in rows:
            ws.append([_stringify(r.get(c, "")) for c in COLUMNS])
            row_idx = ws.max_row
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap
            ci = COLUMNS.index("veredicto_umbral") + 1
            cell = ws.cell(row=row_idx, column=ci)

            # Determinar color según las dos métricas
            try:
                coinc = r.get("coincidencia_%")
                rel   = r.get("relevancia_%")
                coinc_v = float(coinc) if coinc is not None and coinc != "" else None
                rel_v   = float(rel)   if rel   is not None and rel   != "" else None
            except (TypeError, ValueError):
                coinc_v = rel_v = None

            verdict = str(cell.value).upper()
            if verdict == "OK":
                # OK siempre implica ambas >= umbral (65). Verde intenso si ambas >= 75.
                if coinc_v is not None and rel_v is not None and coinc_v >= 75 and rel_v >= 75:
                    cell.fill = ok_dark_fill
                else:
                    cell.fill = ok_light_fill
            elif verdict == "KO":
                # Amarillo si al menos una métrica >= 65, rojo si ninguna lo supera.
                any_above = (
                    (coinc_v is not None and coinc_v >= 65)
                    or (rel_v is not None and rel_v >= 65)
                )
                cell.fill = ko_yellow_fill if any_above else ko_red_fill

        widths = {
            "id": 6, "categoria": 18, "pregunta": 50, "respuesta_esperada": 60,
            "respuesta": 60, "num_llamadas_llm": 8, "tiempo_total_s": 10,
            "tokens_in": 10, "tokens_out": 10, "coincidencia_%": 12,
            "relevancia_%": 12, "veredicto_umbral": 14,
            "justificacion": 50,
            "coste_rag_usd": 14, "coste_eval_usd": 14, "coste_total_usd": 15,
            "error": 30,
        }
        for i, col in enumerate(COLUMNS, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(col, 15)
        ws.freeze_panes = "A2"

    out_path = EVAL_DIR / f"{use_case}.xlsx"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Núcleo
# ---------------------------------------------------------------------------
def evaluate_results_file(
    results_path: Path,
    model_name: str,
    threshold: int,
    batch_size: int = BATCH_SIZE,
) -> tuple[str, str, list[dict[str, Any]], dict[str, int]]:
    """Evalúa un fichero de resultados por lotes.

    Devuelve ``(use_case_rag, idioma, filas, contadores)``.
    """
    with results_path.open(encoding="utf-8") as fh:
        data = json.load(fh)

    use_case = data.get("use_case_rag") or "unknown"
    idioma   = data.get("idioma") or "??"
    preguntas = data.get("preguntas", [])

    print(f"\n📊 Evaluando {results_path.name}  "
          f"(use_case={use_case}, lang={idioma}, n={len(preguntas)}, "
          f"batch_size={batch_size})")

    filas: list[dict[str, Any]] = []
    counters = {"OK_umbral": 0, "KO_umbral": 0,
                "errores_eval": 0, "errores_rag": 0,
                "coste_total_usd": 0.0}

    # Separar preguntas con error RAG vs evaluables
    evaluable_indices: list[int] = []
    for idx, q in enumerate(preguntas):
        if q.get("error"):
            counters["errores_rag"] += 1
            counters["KO_umbral"] += 1
        else:
            evaluable_indices.append(idx)

    # --- Coincidencia + relevancia por lotes (LLM, en paralelo) ---
    eval_workers = _env_int("EVAL_WORKERS", 4)
    print(f"   🤖 Evaluando coincidencia y relevancia (LLM) en lotes de {batch_size} "
          f"(workers={eval_workers})…")

    # Construir todos los lotes de antemano para poder paralelizar.
    batches: list[list[dict[str, Any]]] = []
    for batch_start in range(0, len(evaluable_indices), batch_size):
        batch_end = min(batch_start + batch_size, len(evaluable_indices))
        batch_items = []
        for i in range(batch_start, batch_end):
            idx = evaluable_indices[i]
            q = preguntas[idx]
            batch_items.append({
                "pregunta": q.get("pregunta", ""),
                "esperada": q.get("respuesta_esperada", ""),
                "real": q.get("respuesta", ""),
            })
        batches.append(batch_items)

    total_batches = len(batches)
    coincidencia_results: list[dict[str, Any]] = [None] * len(evaluable_indices)  # type: ignore

    if eval_workers <= 1 or total_batches <= 1:
        for bi, batch_items in enumerate(batches, 1):
            res = evaluate_batch_metrics(batch_items, model_name)
            offset = (bi - 1) * batch_size
            for j, r in enumerate(res):
                coincidencia_results[offset + j] = r
            print(f"      batch {bi}/{total_batches} OK")
    else:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        done = 0
        with ThreadPoolExecutor(max_workers=eval_workers) as ex:
            futures = {
                ex.submit(evaluate_batch_metrics, b, model_name): bi
                for bi, b in enumerate(batches)
            }
            for fut in as_completed(futures):
                bi = futures[fut]
                try:
                    res = fut.result()
                except Exception as exc:
                    res = [{
                        "coincidencia_pct": None, "relevancia_pct": None,
                        "justificacion": "",
                        "eval_error": f"{type(exc).__name__}: {exc}",
                        "tokens_eval_in": 0, "tokens_eval_out": 0,
                    } for _ in batches[bi]]
                offset = bi * batch_size
                for j, r in enumerate(res):
                    coincidencia_results[offset + j] = r
                done += 1
                print(f"      batch {done}/{total_batches} OK (idx={bi+1})")

    # --- Construir filas ---
    eval_counter = 0
    for idx, q in enumerate(preguntas):
        if q.get("error"):
            row = {
                "id": q.get("id"),
                "categoria": q.get("categoria"),
                "pregunta": q.get("pregunta"),
                "respuesta_esperada": q.get("respuesta_esperada"),
                "respuesta": q.get("respuesta"),
                "num_llamadas_llm": q.get("num_llamadas_llm"),
                "tiempo_total_s": q.get("tiempo_total_s"),
                "tokens_in": q.get("tokens_in"),
                "tokens_out": q.get("tokens_out"),
                "coincidencia_%": None,
                "relevancia_%":   None,
                "veredicto_umbral": "KO",
                "justificacion": "Pregunta no evaluada: el RAG falló.",
                "coste_rag_usd":   round(float(q.get("coste_rag_usd") or 0.0), 8),
                "coste_eval_usd":  0.0,
                "coste_total_usd": round(float(q.get("coste_rag_usd") or 0.0), 8),
                "error": q.get("error"),
            }
            filas.append(row)
            continue

        coinc_result = coincidencia_results[eval_counter]
        rel_score    = coinc_result.get('relevancia_pct')
        eval_counter += 1

        coinc = coinc_result["coincidencia_pct"]
        if coinc is not None and rel_score is not None:
            umbral = "OK" if (coinc >= threshold and rel_score >= threshold) else "KO"
        else:
            umbral = "KO"

        counters["OK_umbral" if umbral == "OK" else "KO_umbral"] += 1
        if coinc_result.get("eval_error"):
            counters["errores_eval"] += 1

        justif = coinc_result["justificacion"]
        if coinc_result.get("eval_error"):
            justif += f" [eval_error: {coinc_result['eval_error']}]"

        # --- Cálculo de costes ---
        coste_rag  = round(float(q.get("coste_rag_usd") or 0.0), 8)
        coste_eval = round(
            token_cost_usd(
                model_name,
                coinc_result.get("tokens_eval_in",  0),
                coinc_result.get("tokens_eval_out", 0),
            ),
            8,
        )

        row = {
            "id": q.get("id"),
            "categoria": q.get("categoria"),
            "pregunta": q.get("pregunta"),
            "respuesta_esperada": q.get("respuesta_esperada"),
            "respuesta": q.get("respuesta"),
            "num_llamadas_llm": q.get("num_llamadas_llm"),
            "tiempo_total_s": q.get("tiempo_total_s"),
            "tokens_in": q.get("tokens_in"),
            "tokens_out": q.get("tokens_out"),
            "coincidencia_%": coinc,
            "relevancia_%":   rel_score,
            "veredicto_umbral": umbral,
            "justificacion": justif,
            "coste_rag_usd":   coste_rag,
            "coste_eval_usd":  coste_eval,
            "coste_total_usd": round(coste_rag + coste_eval, 8),
            "error": "",
        }
        filas.append(row)

    # Acumular coste total del fichero
    counters["coste_total_usd"] = round(
        sum(r.get("coste_total_usd") or 0.0 for r in filas), 6
    )

    # Resumen por consola
    total_eval = len(evaluable_indices)
    print(f"   \u2192 OK={counters['OK_umbral']}/{total_eval}, "
          f"KO={counters['KO_umbral']}/{total_eval}, "
          f"errores_eval={counters['errores_eval']}, "
          f"errores_rag={counters['errores_rag']} | "
          f"coste_total=${counters['coste_total_usd']:.6f}")

    return use_case, idioma, filas, counters


def main() -> None:
    _import_openpyxl()  # Verificar que openpyxl está instalado antes de nada
    parser = argparse.ArgumentParser(
        description="Evalúa los resultados del RAG (coincidencia y relevancia con LLM por lotes).",
    )
    parser.add_argument(
        "--model", default=DEFAULT_MODEL,
        help=f"Modelo evaluador (key dentro de MODELS_CONFIG). Default: {DEFAULT_MODEL}",
    )
    parser.add_argument(
        "--threshold", type=int, default=DEFAULT_THRESHOLD,
        help=f"Umbral (%%) para veredicto_umbral. Default: {DEFAULT_THRESHOLD}",
    )
    parser.add_argument(
        "--batch-size", type=int, default=BATCH_SIZE,
        help=f"Preguntas por llamada LLM. Default: {BATCH_SIZE}",
    )
    parser.add_argument(
        "--use-case", help="Filtra por use_case (cvs|eu|wiki).",
    )
    parser.add_argument(
        "--language", help="Filtra por idioma (es|en|fr|it|pt).",
    )
    parser.add_argument(
        "--file", help="Evalúa sólo un fichero results_*.json concreto.",
    )
    args = parser.parse_args()

    if args.file:
        candidate = Path(args.file)
        if not candidate.is_absolute():
            candidate = RESULTS_DIR / candidate.name
        if not candidate.exists():
            sys.exit(f"❌ No se encontró {candidate}")
        files = [candidate]
    else:
        files = sorted(RESULTS_DIR.glob(RESULTS_GLOB))

    if not files:
        sys.exit(f"❌ No hay ficheros de resultados en {RESULTS_DIR}.")

    # Filtrado temprano por nombre de fichero para evitar evaluar resultados
    # que el usuario no ha solicitado (ahorra tiempo y coste).
    if args.use_case or args.language:
        filtered_files = []
        for fp in files:
            m = re.match(r"^results_gold_standard_([a-z]+)_([a-z]{2})\.json$", fp.name)
            if not m:
                continue
            f_use_case, f_lang = m.group(1), m.group(2)
            if args.use_case and f_use_case != args.use_case:
                continue
            if args.language and f_lang != args.language:
                continue
            filtered_files.append(fp)
        files = filtered_files

    if not files:
        sys.exit("❌ No hay ficheros que cumplan los filtros solicitados.")

    by_use_case: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(dict)
    summary: dict[str, dict[str, int]] = {}

    for f in files:
        try:
            use_case, idioma, filas, counters = evaluate_results_file(
                f, model_name=args.model, threshold=args.threshold,
                batch_size=args.batch_size,
            )
        except Exception as exc:
            print(f"❌ Falló evaluación de {f.name}: {exc}")
            traceback.print_exc()
            continue

        by_use_case[use_case][idioma] = filas
        summary[f"{use_case}/{idioma}"] = counters

    if not by_use_case:
        sys.exit("❌ No se generó ningún resultado evaluable.")

    print("\n" + "=" * 70)
    print("📈 Resumen por (use_case/idioma)")
    print("=" * 70)
    grand_total_cost = 0.0
    for key, c in summary.items():
        total = c["OK_umbral"] + c["KO_umbral"]
        grand_total_cost += c.get("coste_total_usd", 0.0)
        print(f"  {key:>14} \u2192 "
              f"OK={c['OK_umbral']}/{total}, KO={c['KO_umbral']}/{total}, "
              f"errores_eval={c['errores_eval']}, errores_rag={c['errores_rag']} | "
              f"coste=${c.get('coste_total_usd', 0.0):.6f}")
    print(f"\n  {'COSTE TOTAL':>14} \u2192 ${grand_total_cost:.6f}")

    print("\n💾 Generando Excels…")
    for use_case, sheets in by_use_case.items():
        out = write_excel_for_use_case(use_case, sheets)
        print(f"   · {out.relative_to(REPO_ROOT)}  ({len(sheets)} hoja(s))")

    print("\n✅ Evaluación finalizada.")


if __name__ == "__main__":
    main()
